# Prompt 1 of this project


"""
EIS API Forge
──────────────────
Upload JSON spec + Excel test cases · Select EIS / NON-EIS ·
LLM mutates payloads · EIS encryption pipeline · Dashboard + export
"""

import streamlit as st
import pandas as pd
import json, time, re, uuid, requests, warnings
from datetime import datetime
from io import BytesIO

warnings.filterwarnings("ignore")  # suppress SSL warnings in console

# ══ Page config ═══════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="EIS API Forge",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ══ CSS ═══════════════════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap');

html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

/* Banner */
.banner {
    background: linear-gradient(135deg, #060d1a 0%, #0d2137 50%, #0f3460 100%);
    border-radius: 14px; padding: 1.5rem 2rem; margin-bottom: 1.4rem;
    display: flex; align-items: center; gap: 1.2rem;
    border: 1px solid #1e3a5f;
}
.banner-icon  { font-size: 2.6rem; line-height: 1; }
.banner-title { font-size: 1.55rem; font-weight: 700; color: #e0f2fe; letter-spacing: -.02em; }
.banner-sub   { font-size: .82rem; color: #7dd3fc; margin-top: .2rem; }

/* Metric tiles */
.metric-row { display: flex; gap: 1rem; margin-bottom: 1.2rem; }
.metric-tile {
    flex: 1; background: #fff; border: 1px solid #e2e8f0;
    border-radius: 12px; padding: 1rem 1.2rem; text-align: center;
    box-shadow: 0 1px 4px rgba(0,0,0,.05);
}
.metric-num   { font-size: 2rem; font-weight: 700; font-family: 'JetBrains Mono', monospace; }
.metric-label { font-size: .68rem; color: #64748b; text-transform: uppercase; letter-spacing: .08em; margin-top: .2rem; }

/* Table badges */
.badge {
    display: inline-block; padding: 3px 13px; border-radius: 20px;
    font-size: .71rem; font-weight: 700;
}
.badge-pass    { background: #dcfce7; color: #16a34a; border: 1px solid #86efac; }
.badge-fail    { background: #fee2e2; color: #dc2626; border: 1px solid #fca5a5; }
.badge-running { background: #dbeafe; color: #2563eb; border: 1px solid #93c5fd; }

/* Code / JSON blocks */
.code-block {
    background: #0d1117; border: 1px solid #30363d; border-radius: 10px;
    padding: .9rem 1rem; font-family: 'JetBrains Mono', monospace;
    font-size: .76rem; color: #79c0ff; white-space: pre-wrap;
    word-break: break-all; max-height: 240px; overflow-y: auto; line-height: 1.6;
}
.code-label {
    font-size: .66rem; font-weight: 700; letter-spacing: .09em;
    text-transform: uppercase; color: #64748b; margin-bottom: .3rem;
}

/* Diff highlights */
.diff-changed { background:#fef9c3; border-left:3px solid #eab308; padding:0 4px; border-radius:3px; }
.diff-added   { background:#dcfce7; border-left:3px solid #22c55e; padding:0 4px; border-radius:3px; }
.diff-removed { background:#fee2e2; border-left:3px solid #dc2626; padding:0 4px; border-radius:3px; }

/* Callouts */
.callout { border-radius: 10px; padding: .85rem 1.1rem; margin: .4rem 0 .8rem; font-size: .82rem; line-height: 1.7; }
.callout-blue  { background:#eff6ff; border:1.5px solid #93c5fd; color:#1e40af; }
.callout-amber { background:#fffbeb; border:1.5px solid #fcd34d; color:#92400e; }
.callout-green { background:#f0fdf4; border:1.5px solid #86efac; color:#14532d; }
.callout-red   { background:#fef2f2; border:1.5px solid #fca5a5; color:#991b1b; }

/* Progress card */
.prog-card {
    background:#fff; border:1px solid #e2e8f0; border-radius:12px;
    padding:1rem 1.3rem; margin-bottom:.8rem;
    box-shadow:0 1px 3px rgba(0,0,0,.05);
}
.prog-card-title { font-size:.75rem; font-weight:700; color:#475569; margin-bottom:.4rem; }

/* Sidebar tweaks */
section[data-testid="stSidebar"] {
    background: #f8fafc !important;
    border-right: 1px solid #e2e8f0 !important;
}
section[data-testid="stSidebar"] * { color: #1e293b !important; }
section[data-testid="stSidebar"] .stMarkdown h3 {
    font-size: .92rem !important; font-weight: 700 !important; color: #0f172a !important;
}
.sb-divider { height:1px; background:#e2e8f0; margin:.6rem 0; }
.sb-section  {
    font-size:.65rem; font-weight:700; letter-spacing:.1em; text-transform:uppercase;
    color:#64748b !important; margin:.6rem 0 .25rem; padding-bottom:.2rem;
    border-bottom:1px solid #e2e8f0;
}

/* Summary chart bar */
.bar-wrap { background:#f1f5f9; border-radius:6px; height:10px; overflow:hidden; margin-top:.3rem; }
.bar-fill-pass { background:#22c55e; height:100%; border-radius:6px; }
.bar-fill-fail { background:#ef4444; height:100%; border-radius:6px; }
</style>
""", unsafe_allow_html=True)


# ══ LLM CLIENT ════════════════════════════════════════════════════════════════

@st.cache_resource
def get_llm_client():
    from openai import AzureOpenAI
    return AzureOpenAI(
        azure_endpoint=AZURE_ENDPOINT,
        api_key=AZURE_API_KEY,
        api_version=AZURE_API_VER,
    )


def _clean_json(raw: str) -> str:
    raw = raw.strip()
    raw = re.sub(r'^```json\s*', '', raw)
    raw = re.sub(r'^```\s*',     '', raw)
    raw = re.sub(r'\s*```$',     '', raw)
    return raw.strip()


# ══ RRN + REGION ══════════════════════════════════════════════════════════════

REGION_DISPLAY = {
    "O":  "O  →  R1",
    "X":  "X  →  R1X",
    "Y":  "Y  →  R1Y",
    "Z":  "Z  →  R1Z",
    "K":  "K  →  R1K",
    "J":  "J  →  R1J",
    "R1": "R1 →  R1R",
    "Y2": "Y2 →  R12",
    "L":  "L  →  R1L",
    "C":  "C  →  R1C",
}

def get_region(code: str) -> str:
    mapping = {
        "O": "R1", "X": "R1X", "Y": "R1Y", "Z": "R1Z",
        "K": "R1K", "J": "R1J", "R1": "R1R", "Y2": "R12",
        "L": "R1L", "C": "R1C",
    }
    return mapping.get(code, "AAA")


def make_rrn(region_code: str) -> str:
    region_val = get_region(region_code)
    uid = uuid.uuid4().hex.upper()[:16]
    return f"SBI{region_val}{uid}"


# ══ EIS CRYPTO ════════════════════════════════════════════════════════════════

def load_eis_keys():
    """Load RSA keys from disk. Returns (private_key, eis_public_key) or (None, None)."""
    try:
        import os
        from Crypto.PublicKey import RSA
        if not os.path.exists(PRIVATE_KEY_PATH) or not os.path.exists(EIS_PUBLIC_KEY_PATH):
            return None, None
        with open(PRIVATE_KEY_PATH, 'r') as f:
            priv = RSA.import_key(f.read())
        with open(EIS_PUBLIC_KEY_PATH, 'r') as f:
            pub  = RSA.import_key(f.read())
        return priv, pub
    except Exception:
        return None, None


def encrypt_aes_gcm_base64(plaintext: str, key: bytes) -> str:
    import base64
    from Crypto.Cipher import AES
    nonce  = key[:12]
    cipher = AES.new(key, AES.MODE_GCM, nonce=nonce)
    ciphertext, tag = cipher.encrypt_and_digest(plaintext.encode('utf-8'))
    return base64.b64encode(ciphertext + tag).decode('utf-8')


def decrypt_aes_gcm_base64(encrypted_b64: str, key: bytes) -> str:
    import base64
    from Crypto.Cipher import AES
    decoded = base64.b64decode(encrypted_b64)
    nonce   = key[:12]
    ct, tag = decoded[:-16], decoded[-16:]
    cipher  = AES.new(key, AES.MODE_GCM, nonce=nonce)
    return cipher.decrypt_and_verify(ct, tag).decode('utf-8')


def get_ds(payload_str: str, private_key) -> str:
    import base64
    from Crypto.Hash      import SHA256
    from Crypto.Signature import pkcs1_15
    h   = SHA256.new(payload_str.encode('utf-8'))
    sig = pkcs1_15.new(private_key).sign(h)
    return base64.b64encode(sig).decode('utf-8')


def get_at(eis_public_key, secret_key: bytes) -> str:
    import base64
    from Crypto.Cipher import PKCS1_OAEP
    enc = PKCS1_OAEP.new(eis_public_key).encrypt(secret_key)
    return base64.b64encode(enc).decode('utf-8')


def get_data_from_eis(payload_str: str, rrn: str, url: str,
                      private_key, eis_public_key) -> tuple[int, dict]:
    """Full EIS call: encrypt → POST → decrypt → verify signature."""
    import base64
    from Crypto.Hash      import SHA256
    from Crypto.Signature import pkcs1_15

    SECRET_KEY = b'11111111111111111111111111111111'
    encrypted  = encrypt_aes_gcm_base64(payload_str, SECRET_KEY)
    d_s        = get_ds(payload_str, private_key)
    a_t        = get_at(eis_public_key, SECRET_KEY)

    headers = {
        'Content-Type': 'application/json',
        'AccessToken':  a_t,
    }
    body = {
        "DIGI_SIGN":                d_s,
        "REQUEST":                  encrypted,
        "REQUEST_REFERENCE_NUMBER": rrn,
    }

    try:
        r = requests.post(url, headers=headers, json=body, verify=False, timeout=60)
        raw = r.json()

        decrypted = decrypt_aes_gcm_base64(raw['RESPONSE'], SECRET_KEY)

        decoded_sign = base64.b64decode(raw['DIGI_SIGN'])
        h = SHA256.new(decrypted.encode('utf-8'))
        try:
            pkcs1_15.new(eis_public_key).verify(h, decoded_sign)
        except Exception:
            pass  # log but don't crash the test run

        try:
            return r.status_code, json.loads(decrypted)
        except Exception:
            return r.status_code, {"raw": decrypted}

    except requests.exceptions.Timeout:
        return 0, {"error": "Request timed out"}
    except Exception as e:
        return 0, {"error": str(e)}


# ══ NON-EIS CALL ══════════════════════════════════════════════════════════════

def call_non_eis(url: str, payload: dict) -> tuple[int, dict]:
    try:
        r = requests.post(
            url, json=payload,
            headers={"Content-Type": "application/json"},
            verify=False, timeout=60,
        )
        try:
            return r.status_code, r.json()
        except Exception:
            return r.status_code, {"raw": r.text[:600]}
    except requests.exceptions.Timeout:
        return 0, {"error": "Request timed out"}
    except Exception as e:
        return 0, {"error": str(e)}


# ══ SIMULATE ══════════════════════════════════════════════════════════════════

def simulate_response(test_case: dict, payload: dict, sample_response: dict) -> tuple[int, dict]:
    tc_type = str(test_case.get("Positive/ Negative", "")).strip().upper()
    prompt  = f"""Simulate an API response.

Sample success response:
{json.dumps(sample_response, indent=2)}

Test case description: {test_case.get('Test Case Description', '')}
Test type            : {tc_type}
Payload sent         : {json.dumps(payload)[:500]}

Rules:
- POSITIVE → HTTP 200, realistic success response matching sample structure.
- NEGATIVE → appropriate error code (400/401/422/500) + error body with ERROR_CODE and ERROR_DESCRIPTION.

Return ONLY: {{"http_code": 200, "body": {{...}}}}"""

    try:
        resp = get_llm_client().chat.completions.create(
            model=AZURE_MODEL,
            messages=[
                {"role": "system", "content": "Simulate API responses. Return only JSON."},
                {"role": "user",   "content": prompt},
            ],
            temperature=0.0, max_tokens=500,
        )
        data = json.loads(_clean_json(resp.choices[0].message.content))
        return int(data.get("http_code", 200)), data.get("body", {})
    except Exception:
        if tc_type == "POSITIVE":
            return 200, {**sample_response, "_simulated": True}
        return 400, {"ERROR_CODE": "400", "ERROR_DESCRIPTION": "Simulated error", "_simulated": True}


# ══ LLM PAYLOAD BUILDER ═══════════════════════════════════════════════════════

def llm_build_payload(test_case: dict, baseline: dict) -> dict:
    tc_id    = test_case.get("Test Case ID", "TC_?")
    tc_desc  = test_case.get("Test Case Description", "")
    tc_type  = str(test_case.get("Positive/ Negative", "")).strip().upper()
    expected = test_case.get("Expected Result", "")

    prompt = f"""You are a QA engineer. Build the exact HTTP request payload for this test case.

Baseline (valid) payload:
{json.dumps(baseline, indent=2)}

Test Case ID  : {tc_id}
Description   : {tc_desc}
Type          : {tc_type}
Expected      : {expected}

Mutation rules — start from baseline, apply ONE mutation based on description:
- POSITIVE              → use baseline unchanged
- blank/empty field     → set that field to ""
- null field            → set that field to null
- invalid field         → set that field to "INVALID_XYZ"
- missing field         → remove that field entirely
- wrong format          → set field to wrong type value
- special characters    → set field to "!@#$%^&*()"
- exceeds max length    → set field to 200+ char string

Do NOT include REQUEST_REFERENCE_NUMBER in the output payload — it is managed separately.

Return ONLY the JSON payload object. No markdown, no explanation."""

    try:
        resp = get_llm_client().chat.completions.create(
            model=AZURE_MODEL,
            messages=[
                {"role": "system", "content": "Return only a JSON payload. No markdown, no text."},
                {"role": "user",   "content": prompt},
            ],
            temperature=0.0, max_tokens=700,
        )
        result = json.loads(_clean_json(resp.choices[0].message.content))
        # strip RRN if LLM accidentally included it
        result.pop("REQUEST_REFERENCE_NUMBER", None)
        return result
    except Exception as e:
        clean = {k: v for k, v in baseline.items() if k != "REQUEST_REFERENCE_NUMBER"}
        clean["_llm_error"] = str(e)
        return clean


# ══ LLM VERDICT ═══════════════════════════════════════════════════════════════

def llm_verdict(test_case: dict, http_code: int, response: dict) -> tuple[str, str]:
    tc_type  = str(test_case.get("Positive/ Negative", "")).strip().upper()
    expected = test_case.get("Expected Result", "")

    prompt = f"""PASS or FAIL for this API test case?

Description  : {test_case.get('Test Case Description', '')}
Expected     : {expected}
Type         : {tc_type}
HTTP Code    : {http_code}
Response     : {json.dumps(response)[:800]}

Rules:
- POSITIVE PASSES when HTTP 200 AND response has no error indicators.
- NEGATIVE PASSES when response has error OR HTTP ≠ 200.
- Check ERROR_CODE, ERROR_DESCRIPTION, RESPONSE_STATUS ≠ "0".

Return ONLY: {{"verdict":"PASS","note":"one sentence"}} or {{"verdict":"FAIL","note":"one sentence"}}"""

    try:
        resp = get_llm_client().chat.completions.create(
            model=AZURE_MODEL,
            messages=[
                {"role": "system", "content": "QA verdict engine. Output only JSON."},
                {"role": "user",   "content": prompt},
            ],
            temperature=0.0, max_tokens=80,
        )
        d = json.loads(_clean_json(resp.choices[0].message.content))
        return d.get("verdict", "FAIL"), d.get("note", "")
    except Exception:
        if tc_type == "POSITIVE":
            v = "PASS" if http_code == 200 else "FAIL"
        else:
            s   = json.dumps(response).lower()
            err = any(k in s for k in ["error", "fail", "invalid"])
            v   = "PASS" if (http_code != 200 or err) else "FAIL"
        return v, "Heuristic verdict."


# ══ LLM SUMMARY ═══════════════════════════════════════════════════════════════

def llm_summary(results: list[dict]) -> str:
    total  = len(results)
    passed = sum(1 for r in results if r.get("Status") == "PASS")
    failed = total - passed
    tc_lines = "\n".join(
        f"- {r.get('Test Case ID','?')} [{r.get('Positive/ Negative','')}] "
        f"HTTP {r.get('HTTP Code','')} → {r.get('Status','?')}: "
        f"{str(r.get('Test Case Description',''))[:80]}"
        for r in results
    )

    prompt = f"""Write a concise QA test run summary (4-6 sentences, professional tone).

Run stats: {total} total · {passed} PASS · {failed} FAIL

Test results:
{tc_lines}

Cover: overall pass rate, any patterns in failures, notable observations, recommended next steps.
No bullet points. No markdown headers. Plain paragraph."""

    try:
        resp = get_llm_client().chat.completions.create(
            model=AZURE_MODEL,
            messages=[
                {"role": "system", "content": "Professional QA analyst. Write concise summaries."},
                {"role": "user",   "content": prompt},
            ],
            temperature=0.2, max_tokens=300,
        )
        return resp.choices[0].message.content.strip()
    except Exception as e:
        return f"Summary unavailable: {e}"


# ══ DIFF RENDERER ═════════════════════════════════════════════════════════════

def diff_html(baseline: dict, generated: dict) -> str:
    lines = []
    for k, v in generated.items():
        vs       = json.dumps(v)
        base_val = baseline.get(k, "__NEW__")
        if base_val == "__NEW__":
            lines.append(f'  <span class="diff-added">"{k}": {vs}  ← new</span>')
        elif v != base_val:
            lines.append(f'  <span class="diff-changed">"{k}": {vs}  ← was {json.dumps(base_val)}</span>')
        else:
            lines.append(f'  "{k}": {vs}')
    for k in baseline:
        if k not in generated:
            lines.append(f'  <span class="diff-removed">"{k}": {json.dumps(baseline[k])}  ← removed</span>')
    return "{{\n{}\n}}".format(",\n".join(lines))


# ══ EXCEL EXPORT ══════════════════════════════════════════════════════════════

def to_excel(df: pd.DataFrame) -> bytes:
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Results")
        ws = writer.sheets["Results"]
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(
                len(str(col[0].value or "")), 18)
        from openpyxl.styles import PatternFill, Font
        for cell in ws[1]:
            cell.font = Font(bold=True)
        green = PatternFill("solid", fgColor="C6EFCE")
        red   = PatternFill("solid", fgColor="FFC7CE")
        sc    = next((c.column for c in ws[1] if c.value == "Status"), None)
        if sc:
            for row in ws.iter_rows(min_row=2):
                sv   = ws.cell(row[0].row, sc).value
                fill = green if sv == "PASS" else red if sv == "FAIL" else None
                if fill:
                    for cell in row:
                        cell.fill = fill
    return buf.getvalue()


# ══ MAIN ══════════════════════════════════════════════════════════════════════

def main():

    # ── Session defaults ──────────────────────────────────────────────────────
    for k, v in [
        ("spec",        {}),
        ("df_tc",       pd.DataFrame()),
        ("df_results",  pd.DataFrame()),
        ("llm_summary", ""),
        ("run_done",    False),
    ]:
        if k not in st.session_state:
            st.session_state[k] = v

    # ══════════════════════════════════════════════════════════════════════════
    # SIDEBAR
    # ══════════════════════════════════════════════════════════════════════════
    with st.sidebar:
        st.markdown("### ⚡ EIS API Forge")
        st.markdown('<div class="sb-divider"></div>', unsafe_allow_html=True)

        # ── 1. JSON spec upload ───────────────────────────────────────────────
        st.markdown('<div class="sb-section">1 · API Spec</div>', unsafe_allow_html=True)
        st.caption("JSON file with `api_url`, `method`, `payload`, `sample_response`.")

        json_file = st.file_uploader(
            "API spec (.json)", type=["json"],
            key="json_up", label_visibility="collapsed",
        )
        if json_file:
            try:
                spec = json.load(json_file)
                st.session_state["spec"] = spec
                st.success(f"✅ Spec loaded · `{spec.get('api_url','?')[:40]}`")
            except Exception as e:
                st.error(f"JSON parse error: {e}")

        if st.session_state["spec"]:
            with st.expander("View spec", expanded=False):
                st.json(st.session_state["spec"])

        # ── 2. Excel upload ───────────────────────────────────────────────────
        st.markdown('<div class="sb-divider"></div>', unsafe_allow_html=True)
        st.markdown('<div class="sb-section">2 · Test Cases</div>', unsafe_allow_html=True)
        st.caption("Excel with: Test Case ID, Test Case Description, Expected Result, Positive/ Negative")

        xlsx_file = st.file_uploader(
            "Test cases (.xlsx)", type=["xlsx", "xls"],
            key="xlsx_up", label_visibility="collapsed",
        )
        if xlsx_file:
            try:
                xls    = pd.ExcelFile(xlsx_file)
                sheet  = xls.sheet_names[0]
                df_tc  = pd.read_excel(xls, sheet_name=sheet)
                df_tc.columns = [str(c).strip() for c in df_tc.columns]
                st.session_state["df_tc"] = df_tc
                st.success(f"✅ {len(df_tc)} test cases loaded")
            except Exception as e:
                st.error(f"Excel error: {e}")

        # ── 3. API mode ───────────────────────────────────────────────────────
        st.markdown('<div class="sb-divider"></div>', unsafe_allow_html=True)
        st.markdown('<div class="sb-section">3 · API Mode</div>', unsafe_allow_html=True)

        api_mode = st.radio(
            "Select mode", ["EIS", "NON EIS"],
            horizontal=True, key="api_mode",
            label_visibility="collapsed",
        )

        region_code = None
        if api_mode == "EIS":
            region_display = st.selectbox(
                "Region",
                options=list(REGION_DISPLAY.keys()),
                format_func=lambda k: REGION_DISPLAY[k],
                key="region_sel",
            )
            region_code = region_display
            st.markdown(
                f"<div style='background:#eff6ff;border:1px solid #93c5fd;border-radius:6px;"
                f"padding:.4rem .7rem;font-size:.75rem;color:#1e40af;margin:.3rem 0;'>"
                f"RRN prefix: <b>SBI{get_region(region_code)}…</b></div>",
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                "<div style='background:#f0fdf4;border:1px solid #86efac;border-radius:6px;"
                "padding:.4rem .7rem;font-size:.75rem;color:#14532d;margin:.3rem 0;'>"
                "RRN prefix: <b>SBIAAA…</b></div>",
                unsafe_allow_html=True,
            )
